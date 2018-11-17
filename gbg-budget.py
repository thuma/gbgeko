#!/usr/bin/env python
# -*- coding: utf-8 -*-

import requests
from openpyxl import load_workbook
import sqlite3
conn = sqlite3.connect('gbg.db')

headers = [
  "Förvaltning",
  "Leverantör",
  "Organisationsnummer",
  "Verifikationsnummer",
  "Konto",
  "Kontotext",
  "Belopp"]

types = [
  "text",
  "text",
  "int",
  "int",
  "int",
  "text",
  "real"]

urls = [
  "https://catalog.goteborg.se/store/6/resource/504",
  "https://catalog.goteborg.se/store/6/resource/456",
  "https://catalog.goteborg.se/store/6/resource/346",
  "https://catalog.goteborg.se/store/6/resource/300",
  "https://catalog.goteborg.se/store/6/resource/298",
  "https://catalog.goteborg.se/store/6/resource/293",
  "https://catalog.goteborg.se/store/6/resource/291",
  "https://catalog.goteborg.se/store/6/resource/207",
  "https://catalog.goteborg.se/store/6/resource/175",
  "https://catalog.goteborg.se/store/6/resource/149",
  "https://catalog.goteborg.se/store/6/resource/143",
  "https://catalog.goteborg.se/store/6/resource/106"]

cols = []
for i in range(0,len(headers)):
  cols.append(headers[i] + " " + types[i])
cols = ', '.join(cols)

make = "CREATE TABLE IF NOT EXISTS bills (%s)" % cols
conn.execute(make)
conn.execute('''CREATE UNIQUE INDEX IF NOT EXISTS Verifikationsnummer on bills (Verifikationsnummer)''')
conn.commit()

for url in urls:
  result = requests.get(url);
  with  open('test.xlsx','w') as file:
   file.write(result.content)
  wb2 = load_workbook('test.xlsx')
  for row in wb2[wb2.get_sheet_names()[0]].iter_rows(row_offset=1):
    data = []
    for col in range(0,len(headers)):
      coldata = row[col].value;
      if not coldata:
        coldata = ""
      if isinstance(coldata, unicode):
        coldata = coldata.encode('utf-8')
      coldata = "%s" % coldata
      coldata = coldata.replace('"',"'")
      data.append('"'+coldata+'"')
    data = ','.join(data)
    insert = ("INSERT OR IGNORE INTO bills VALUES (%s)" % data)
    print insert
    conn.execute(insert)
  conn.commit()
